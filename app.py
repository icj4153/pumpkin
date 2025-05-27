from flask import Flask, request, send_file, render_template_string
import pandas as pd
import openpyxl
from io import BytesIO
import datetime

app = Flask(__name__)

HTML_TEMPLATE = """
<!doctype html>
<html lang=\"ko\">
  <head>
    <meta charset=\"utf-8\">
    <title>호박잎 발주서 변환기</title>
  </head>
  <body style=\"font-family: sans-serif; text-align: center; margin-top: 50px;\">
    <h1>🍃 호박잎 공급처용 발주서 변환기</h1>
    <form action=\"/convert\" method=\"post\" enctype=\"multipart/form-data\">
      <label>📥 Delivery_List 엑셀 파일 업로드 (.xlsx)</label><br><br>
      <input type=\"file\" name=\"delivery_file\" accept=\".xlsx\" required><br><br>
      <label>📄 발주서 양식 파일 업로드 (.xlsx)</label><br><br>
      <input type=\"file\" name=\"template_file\" accept=\".xlsx\" required><br><br>
      <button type=\"submit\" style=\"font-size: 16px;\">발주서 만들기</button>
    </form>
  </body>
</html>
"""

COLUMN_MAP = {
    "수취인이름": "받는사람 이름",
    "수취인전화번호": "전화번호",
    "수취인 주소": "받는분주소",
    "구매수(수량)": "수량",
    "등록옵션명": "상품명",
    "구매자":"보내는분 이름",
    "배송메세지": "배송메모"
}

@app.route("/", methods=["GET"])
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route("/convert", methods=["POST"])
def convert():
    delivery_file = request.files['delivery_file']
    template_file = request.files['template_file']

    delivery_df = pd.read_excel(delivery_file)
    template_wb = openpyxl.load_workbook(template_file)
    template_ws = template_wb.active

    # 열 인덱스 찾기 (기존 열 제목 → 열 번호)
    header_row = 1
    col_index = {}
    for col in range(1, template_ws.max_column + 1):
        value = str(template_ws.cell(row=header_row, column=col).value).strip()
        if value in COLUMN_MAP.values():
            col_index[value] = col

    # Delivery_List의 각 행을 템플릿에 채워넣기
    start_row = header_row + 1
    for i, row in delivery_df.iterrows():
        for source_col, target_col in COLUMN_MAP.items():
            if target_col in col_index:
                value = row.get(source_col, "")
                template_ws.cell(row=start_row + i, column=col_index[target_col]).value = value

    # 저장
    output = BytesIO()
    today = datetime.datetime.now().strftime("%y%m%d")
    filename = f"호박잎_발주서_{today}.xlsx"
    template_wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
